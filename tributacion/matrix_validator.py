"""matrix_validator.py — Validación estructural de la Matriz de Tributación.

Antes de que :func:`~tributacion.matrix_parser.parse_matrix` intente extraer
datos, este módulo verifica que el archivo Excel tenga exactamente la forma
que el parser espera. Falla rápido con mensajes claros en lugar de producir
resultados silenciosamente incorrectos.

Contrato estructural de la Matriz de Tributación
-------------------------------------------------
El archivo Excel debe cumplir **todas** estas condiciones:

1. **Hoja correcta**: existe una hoja llamada exactamente ``"Asignaturas - RA"``
   (definida en :data:`~tributacion.config.DEFAULT_SHEET_NAME`).

2. **Dimensiones mínimas**: al menos 6 filas (5 de cabecera + al menos 1
   asignatura) y al menos 4 columnas (A–C de metadata + ≥1 área).

3. **Columna A (index 0), filas 1–5 (rows 0–4)**: celda fusionada con el
   texto ``"ÁREA DE FORMACIÓN"``; es decir, ``df.iloc[0,0]`` tiene ese valor
   y las filas 1–4 deben estar vacías (NaN).

4. **Columna B (index 1), filas 1–5 (rows 0–4)**: celda fusionada con el
   texto ``"SEMESTRE"``; mismo criterio que A.

5. **Columna C (index 2), filas 1–5 (rows 0–4)**: las 5 etiquetas fijas de
   cabecera (``"ÁREA DE FORMACIÓN"``, ``"ÁMBITOS DE REALIZACIÓN"``, etc.).

6. **Fila 1 — Orden de áreas (row index 0)**: a partir de la columna D
   (index 3), los únicos valores no vacíos permitidos son los siguientes títulos
   (en este orden relativo; los opcionales pueden estar ausentes)::

       ÁREA DE FORMACIÓN DISCIPLINAR          (requerido)
       ÁREA DE FORMACIÓN BÁSICA               (opcional, máx. 1 celda)
       ÁREA DE FORMACIÓN ESPECIALIZADA        (opcional, máx. 1 celda)
       ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN INTEGRAL  (requerido)
       MODELO EDUCATIVO INSTITUCIONAL         (requerido)

7. **Fila 1 — Sin huecos entre secciones (row index 0)**: no debe haber
   columnas vacías sin fusionar entre la primera y la última sección de área
   reconocida. Cada columna de ese rango debe pertenecer a alguna celda
   fusionada de encabezado.

7. **Fila 2 — AR (row index 1)**: para las columnas que caen bajo el encabezado
   ``"ÁREA DE FORMACIÓN DISCIPLINAR"`` (determinado por forward-fill de la fila 1),
   cada celda no vacía debe coincidir con **uno** de estos dos patrones:

   - Forma 1: ``AR_<número>: <descripción larga>``
   - Forma 2: ``AR_<número> <nombre corto>: <descripción larga>``

   Los números deben ser **correlativos** de izquierda a derecha: cada celda puede
   repetir el número de la anterior (celda fusionada en el original) o avanzar
   exactamente en 1 respecto a la anterior. El primer número encontrado debe ser 1.

8. **Fila 3 — RA (row index 2)**: para las columnas bajo
   ``"ÁREA DE FORMACIÓN DISCIPLINAR"`` (forward-fill de fila 1) se aplican
   tres sub-reglas:

   a. **Formato obligatorio**: cada celda no vacía debe seguir el patrón
      ``AR_x – RAy: <descripción larga>`` (se acepta guion corto, medio o largo).
   b. **Consistencia de AR**: el número ``x`` de ``AR_x`` debe coincidir con
      el número del ``AR_x`` de la celda directamente arriba en la fila 2
      (Ámbitos de Realización, después de forward-fill).
   c. **Correlatividad de RA**: los valores de ``y`` deben ser un entero
      correlativo global de izquierda a derecha: el primero debe ser ``1``,
      y cada celda posterior puede repetir el número anterior (celda fusionada)
      o avanzar exactamente en 1.

   Adicionalmente, en toda la fila 3 (incluidas columnas fuera de DISCIPLINAR)
   debe haber al menos una celda que contenga ``"–"``, ``"—"`` o ``":"``.

   Para las columnas bajo ``"ÁREA DE FORMACIÓN BÁSICA"`` cada celda no vacía
   debe seguir el patrón ``FB_RAx: <descripción>``
   (o ``FB_RAx: <desc_corta>: <desc_larga>``), con ``x`` correlativo desde 1.

   Para las columnas bajo ``"ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN
   INTEGRAL"`` (o su alias ``"HABILIDADES TRANSVERSALES"`` en fila 1), cada
   celda no vacía que no sea el sub-encabezado ``"HABILIDADES TRANSVERSALES"``
   debe seguir el patrón ``FG_RAx: <descripción>``, con ``x`` correlativo.

9. **Fila 4 — Niveles de logro (row index 3)**: para las columnas bajo
   DISCIPLINAR, BÁSICA y GENERAL/EJE (excluidas las columnas cuya fila 3 es
   ``"HABILIDADES TRANSVERSALES"``), cada celda no vacía debe coincidir con::

       N. INICIAL: <descripción>
       N. INTERMEDIO: <descripción>
       N. AVANZADO: <descripción>   (también se acepta N. TITULACION)

   y seguir ese ciclo correlativo de tres niveles de izquierda a derecha
   (una celda puede repetir el nivel anterior por fusión o avanzar al siguiente).

10. **Columna C (index 2) desde fila 6 (row index 5)**: al menos una celda
    no vacía que represente el nombre de una asignatura.

11. **Columna A (index 0) desde fila 6**: al menos una celda con texto
    (área de formación).

12. **Columna B (index 1) desde fila 6**: al menos una celda con numeral
    romano (I, II, III … X).

13. **Al menos un "1" de tributación**: en las filas de asignatura (≥ fila 7),
    else:
        print("Estructura válida.")
"""

import os
import unicodedata
from pathlib import Path

import pandas as pd

from tributacion.config import (
    AREA_TITLES,
    DEFAULT_SHEET_NAME,
    TRIBUTACION_VALUE,
)

# Numerales romanos aceptados como semestre (I..XII)
_ROMAN_NUMERALS = {"I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"}

# Valores de celda considerados vacíos en comprobaciones de texto
_EMPTY = ("", "nan", "NAN", "NaN")


def _resolve_path(path: Path) -> Path:
    """Resolve a path that may fail due to NFC/NFD Unicode normalization differences.

    On Linux filesystems (ext4, etc.) filenames are stored as raw bytes without
    normalization, so a file saved on macOS (which uses NFD) may not be found
    by Python's ``Path.exists()`` when the caller provides an NFC-encoded path.
    This helper scans each directory component using ``os.scandir`` and compares
    filenames after normalizing both sides to NFC.

    Args:
        path: Path (absolute or relative, NFC or NFD) to resolve.

    Returns:
        The resolved ``Path`` if found, otherwise the original ``path``.
    """
    if path.exists():
        return path

    # Build the path component-by-component using scandir for each part
    parts = path.parts  # e.g. ('data', 'data_raw', 'Matrices...', 'file.xlsx')
    resolved = Path(parts[0])
    for part in parts[1:]:
        part_nfc = unicodedata.normalize("NFC", part)
        try:
            matched = next(
                (e.name for e in os.scandir(str(resolved))
                 if unicodedata.normalize("NFC", e.name) == part_nfc),
                None,
            )
        except OSError:
            matched = None
        resolved = resolved / (matched if matched else part)

    return resolved if resolved.exists() else path


# ---------------------------------------------------------------------------
# Checks privados — uno por condición del contrato estructural
# ---------------------------------------------------------------------------

def _check_file_exists(xlsx_path: Path) -> list[str]:
    """Check 0: el archivo existe en disco (con fallback de normalización Unicode)."""
    if not xlsx_path.exists():
        return [f"El archivo no existe: {xlsx_path}"]
    return []


def _check_sheet_exists(xlsx_path: Path, sheet_name: str) -> list[str]:
    """Check 1: el Excel se puede abrir y la hoja requerida existe."""
    try:
        xl = pd.ExcelFile(str(xlsx_path))
    except Exception as exc:
        return [f"No se pudo abrir el Excel: {exc}"]

    if sheet_name not in xl.sheet_names:
        return [
            f"Hoja '{sheet_name}' no encontrada. "
            f"Hojas disponibles: {xl.sheet_names}"
        ]
    return []


def _check_dimensions(df: pd.DataFrame) -> list[str]:
    """Check 2: la hoja supera el mínimo de filas y columnas."""
    MIN_ROWS = 6   # 5 de cabecera + al menos 1 asignatura
    MIN_COLS = 4   # A(0) B(1) C(2) de metadata + al menos 1 columna de área
    issues: list[str] = []

    if df.shape[0] < MIN_ROWS:
        issues.append(
            f"La hoja tiene solo {df.shape[0]} fila(s); "
            f"se esperan al menos {MIN_ROWS} (5 de cabecera + 1 de datos)."
        )
    if df.shape[1] < MIN_COLS:
        issues.append(
            f"La hoja tiene solo {df.shape[1]} columna(s); "
            f"se esperan al menos {MIN_COLS} (A–C de metadata + ≥1 área)."
        )
    return issues


# Títulos de área obligatorios en fila 1 (row 0), columnas ≥3, en el orden exacto
_REQUIRED_AREA_TITLES_ORDERED: list[str] = [
    "ÁREA DE FORMACIÓN DISCIPLINAR",
    "ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN INTEGRAL",
    "MODELO EDUCATIVO INSTITUCIONAL",
]

# Títulos opcionales (máx. 1 cada uno); el orden relativo esperado es:
#   DISCIPLINAR < BÁSICA < ESPECIALIZADA < GENERAL/EJE < MODELO
_OPTIONAL_AREA_TITLES: list[str] = [
    "ÁREA DE FORMACIÓN BÁSICA",
    "ÁREA DE FORMACIÓN ESPECIALIZADA",
]

# Secuencia global (requeridos + opcionales) que define el orden permitido
_ALL_AREA_TITLES_ORDERED: list[str] = [
    "ÁREA DE FORMACIÓN DISCIPLINAR",
    "ÁREA DE FORMACIÓN BÁSICA",
    "ÁREA DE FORMACIÓN ESPECIALIZADA",
    "ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN INTEGRAL",
    "MODELO EDUCATIVO INSTITUCIONAL",
]

# Títulos extra permitidos en cualquier posición del encabezado (no se reportan como error).
# Incluye columnas de apoyo presentes en matrices reales que no son áreas de tributación.
_KNOWN_EXTRA_TITLES: set[str] = {
    "TOTAL TRIBUTACIÓN",
    "REFERENTES CURRICULARES ESPECÍFICOS POR ASIGNATURA",
    "EN QUÉ COMPONENTE DEL PROGRAMA DE ESTUDIO TRIBUTA (DÓNDE)",
    "HABILIDADES TRANSVERSALES",
}

# Etiquetas exactas esperadas en la columna C (índice 2), filas 0-4
_COL_C_HEADERS = [
    "ÁREA DE FORMACIÓN",
    "ÁMBITOS DE REALIZACIÓN",
    "RESULTADOS DE APRENDIZAJE",
    "NIVELES DE  LOGRO  R.A.",
    "ASIGNATURA",
]


def _check_col_c_headers(df: pd.DataFrame) -> list[str]:
    """Check 2b: la columna C (índice 2), filas 0–4, contiene las 5 etiquetas de cabecera esperadas.

    Estas etiquetas fijas son el indicador más confiable de que el DataFrame
    arranca en la fila correcta y la plantilla no ha sido modificada.
    """
    issues: list[str] = []
    if df.shape[0] < 5:
        return []  # _check_dimensions ya reportará el problema
    for row_idx, expected in enumerate(_COL_C_HEADERS):
        actual = str(df.iloc[row_idx, 2]).strip() if df.shape[1] > 2 else ""
        if actual.upper() != expected.upper():
            issues.append(
                f"Columna C fila {row_idx + 1}: se esperaba '{expected}', "
                f"pero se encontró '{actual}'. "
                "La plantilla tiene filas de preámbulo extra o la estructura no es la esperada."
            )
    return issues


def _check_col_a_header(df: pd.DataFrame) -> list[str]:
    """Check 3a: columna A (índice 0), filas 1–5 (rows 0–4), es celda fusionada 'ÁREA DE FORMACIÓN'.

    Al leer con ``header=None``, las celdas fusionadas dejan el valor en la
    primera posición (row 0) y NaN en las restantes (rows 1–4).
    """
    if df.shape[0] < 5 or df.shape[1] < 1:
        return []
    issues: list[str] = []
    val = str(df.iloc[0, 0]).strip()
    if val.upper() != "ÁREA DE FORMACIÓN":
        issues.append(
            f"Columna A fila 1: se esperaba 'ÁREA DE FORMACIÓN' (celda fusionada "
            f"filas 1–5), pero se encontró '{val}'."
        )
    for r in range(1, 5):
        v = str(df.iloc[r, 0]).strip()
        if v not in _EMPTY:
            issues.append(
                f"Columna A fila {r + 1}: se esperaba celda vacía (parte de la "
                f"fusión con fila 1), pero se encontró '{v}'."
            )
    return issues


def _check_col_b_header(df: pd.DataFrame) -> list[str]:
    """Check 3b: columna B (índice 1), filas 1–5 (rows 0–4), es celda fusionada 'SEMESTRE'.

    Al leer con ``header=None``, las celdas fusionadas dejan el valor en la
    primera posición (row 0) y NaN en las restantes (rows 1–4).
    """
    if df.shape[0] < 5 or df.shape[1] < 2:
        return []
    issues: list[str] = []
    val = str(df.iloc[0, 1]).strip()
    if val.upper() != "SEMESTRE":
        issues.append(
            f"Columna B fila 1: se esperaba 'SEMESTRE' (celda fusionada "
            f"filas 1–5), pero se encontró '{val}'."
        )
    for r in range(1, 5):
        v = str(df.iloc[r, 1]).strip()
        if v not in _EMPTY:
            issues.append(
                f"Columna B fila {r + 1}: se esperaba celda vacía (parte de la "
                f"fusión con fila 1), pero se encontró '{v}'."
            )
    return issues


def _check_area_title_order(df: pd.DataFrame) -> list[str]:
    """Check 3c: fila 1 (row 0) — títulos de área presentes, únicos y en orden.

    Reglas:
    - Requeridos (deben aparecer exactamente una vez): DISCIPLINAR,
      GENERAL/EJE, MODELO EDUCATIVO INSTITUCIONAL.
    - Opcionales (pueden estar ausentes, pero si aparecen, máx. una vez):
      BÁSICA, ESPECIALIZADA.
    - El orden relativo se verifica solo entre DISCIPLINAR, BÁSICA (si existe),
      GENERAL/EJE y MODELO. ESPECIALIZADA puede aparecer en cualquier posición.
    - Los títulos de ``_KNOWN_EXTRA_TITLES`` se ignoran (vienen después).
    - Cualquier otro título es un error.
    """
    if df.shape[0] < 1 or df.shape[1] < 4:
        return []

    # Recopilar valores no vacíos de fila 0 a partir de la col índice 3
    non_empty: list[tuple[int, str]] = [
        (c, str(df.iloc[0, c]).strip())
        for c in range(3, df.shape[1])
        if str(df.iloc[0, c]).strip() not in _EMPTY
    ]

    required_upper = [t.upper() for t in _REQUIRED_AREA_TITLES_ORDERED]
    optional_upper = [t.upper() for t in _OPTIONAL_AREA_TITLES]
    extra_upper = {t.upper() for t in _KNOWN_EXTRA_TITLES}
    allowed_upper = set(required_upper) | set(optional_upper) | extra_upper

    # Títulos cuyo orden relativo se exige: requeridos + BÁSICA (sin ESPECIALIZADA)
    _BASICA_UPPER = "ÁREA DE FORMACIÓN BÁSICA".upper()
    ordered_subset_upper = [t.upper() for t in _ALL_AREA_TITLES_ORDERED
                            if t.upper() != "ÁREA DE FORMACIÓN ESPECIALIZADA".upper()]

    issues: list[str] = []

    # --- Títulos completamente inesperados ---
    unexpected = [t for _, t in non_empty if t.upper() not in allowed_upper]
    if unexpected:
        allowed_display = (
            [f"'{t}'" for t in _REQUIRED_AREA_TITLES_ORDERED]
            + [f"'{t}' (opcional)" for t in _OPTIONAL_AREA_TITLES]
        )
        issues.append(
            f"Fila 1: títulos de área no reconocidos: {unexpected}. "
            "Permitidos: " + ", ".join(allowed_display) + "."
        )

    # --- Duplicados en opcionales ---
    for opt, opt_u in zip(_OPTIONAL_AREA_TITLES, optional_upper):
        count = sum(1 for _, t in non_empty if t.upper() == opt_u)
        if count > 1:
            issues.append(
                f"Fila 1: '{opt}' aparece {count} veces; "
                "solo se permite una única celda (aunque abarque varias columnas)."
            )

    # --- Presencia de requeridos ---
    found_pos: dict[str, int] = {}  # upper → índice en non_empty
    titles_upper_list = [t.upper() for _, t in non_empty]
    for req, req_u in zip(_REQUIRED_AREA_TITLES_ORDERED, required_upper):
        try:
            found_pos[req_u] = titles_upper_list.index(req_u)
        except ValueError:
            issues.append(
                f"Fila 1: falta el título de área requerido '{req}'. "
                f"Títulos encontrados: {[t for _, t in non_empty]}"
            )

    # --- Orden relativo (excluye ESPECIALIZADA) ---
    if len(found_pos) == len(required_upper):
        # Añadir BÁSICA si está presente
        if _BASICA_UPPER in titles_upper_list:
            found_pos[_BASICA_UPPER] = titles_upper_list.index(_BASICA_UPPER)

        present_ordered = [u for u in ordered_subset_upper if u in found_pos]
        actual_positions = [found_pos[u] for u in present_ordered]
        if actual_positions != sorted(actual_positions):
            actual_labels = [(non_empty[found_pos[u]][0], non_empty[found_pos[u]][1])
                             for u in present_ordered]
            expected_labels = [_ALL_AREA_TITLES_ORDERED[
                [t.upper() for t in _ALL_AREA_TITLES_ORDERED].index(u)] for u in present_ordered]
            issues.append(
                "Fila 1: los títulos de área no están en el orden correcto. "
                f"Orden esperado (de los presentes): {expected_labels}. "
                f"Encontrados (col, valor): {actual_labels}"
            )

    return issues


def _check_area_row_gaps(xlsx_path: Path, sheet_name: str) -> list[str]:
    """Check 3d: fila 1 no tiene columnas vacías sin fusionar entre secciones de área.

    Una columna es un "hueco" si:
    - está entre la primera y la última columna de área reconocida (incl. extras), Y
    - no pertenece a ninguna celda fusionada, Y
    - está vacía (``None`` en openpyxl).

    Usa openpyxl para leer la información de celdas fusionadas, que pandas
    no expone al leer con ``header=None``.
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except Exception:
        return []  # si no se puede abrir, _check_sheet_exists ya habrá fallado
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    # Columnas cubiertas por alguna celda fusionada en fila 1
    merged_cols: set[int] = set()
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= 1 <= mr.max_row:
            for c in range(mr.min_col, mr.max_col + 1):
                merged_cols.add(c)

    # Títulos reconocidos (área + extras)
    all_known_upper = (
        {t.upper() for t in _ALL_AREA_TITLES_ORDERED}
        | _KNOWN_EXTRA_TITLES
    )

    # Localizar todas las columnas que contienen un título reconocido en fila 1
    title_cols: list[int] = []
    for col in range(4, (ws.max_column or 0) + 1):
        v = ws.cell(1, col).value
        if v is not None and str(v).strip().upper() in all_known_upper:
            title_cols.append(col)

    if len(title_cols) < 2:
        return []

    first_col = title_cols[0]

    # Extender hasta el final de la celda fusionada del último título
    last_title_col = title_cols[-1]
    last_col = last_title_col
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= 1 <= mr.max_row and mr.min_col == last_title_col:
            last_col = mr.max_col
            break

    # Detectar huecos: vacíos + no fusionados
    issues: list[str] = []
    gap_start: int | None = None

    for col in range(first_col, last_col + 1):
        v = ws.cell(1, col).value
        is_empty = v is None or str(v).strip() == ""
        is_gap = is_empty and col not in merged_cols

        if is_gap:
            if gap_start is None:
                gap_start = col
        else:
            if gap_start is not None:
                gap_end = col - 1
                s = get_column_letter(gap_start)
                e = get_column_letter(gap_end)
                span = f"{s}1" if gap_start == gap_end else f"{s}1:{e}1"
                issues.append(
                    f"Fila 1: columnas {span} están vacías y sin fusionar — "
                    "hueco entre secciones de área. Cada columna del rango de "
                    "encabezados debe pertenecer a alguna celda fusionada de área."
                )
                gap_start = None

    if gap_start is not None:
        gap_end = last_col
        s = get_column_letter(gap_start)
        e = get_column_letter(gap_end)
        span = f"{s}1" if gap_start == gap_end else f"{s}1:{e}1"
        issues.append(
            f"Fila 1: columnas {span} están vacías y sin fusionar — "
            "hueco al final del rango de encabezados de área."
        )

    return issues


def _check_ambitos_realizacion(df: pd.DataFrame) -> list[str]:
    """Check 4: fila 2 (row 1) — bajo 'ÁREA DE FORMACIÓN DISCIPLINAR', formato y consecutividad de AR.

    Reglas:
    - Solo se validan las columnas cuyo encabezado en fila 1 (row 0) corresponde a
      ``'ÁREA DE FORMACIÓN DISCIPLINAR'``, determinando el rango mediante forward-fill.
    - Cada celda **no vacía** en ese rango debe coincidir con **uno** de los dos
      patrones aceptados (insensible a mayúsculas):

        - Forma 1: ``AR_<número>: <descripción larga>``
        - Forma 2: ``AR_<número> <nombre corto>: <descripción larga>``

    - Los números deben ser **correlativos** de izquierda a derecha:
        - El primer número encontrado debe ser 1.
        - Cada celda posterior puede repetir el número anterior (celda fusionada)
          o avanzar exactamente en 1 (``num == prev_num`` o ``num == prev_num + 1``).
    - Si no hay ninguna celda con ese formato bajo DISCIPLINAR, se reporta error.
    """
    import re
    from openpyxl.utils import get_column_letter

    if df.shape[0] < 2 or df.shape[1] < 4:
        return []

    # Forward-fill fila 0 para cubrir las columnas de celdas fusionadas
    area_row_ffill = df.iloc[0].ffill().fillna("").astype(str)

    disciplinar_upper = "ÁREA DE FORMACIÓN DISCIPLINAR".upper()
    disciplinar_cols = [
        c for c in range(3, df.shape[1])
        if area_row_ffill.iloc[c].strip().upper() == disciplinar_upper
    ]

    if not disciplinar_cols:
        # _check_area_title_order ya reportará la ausencia del área requerida
        return []

    ar_row = df.iloc[1].fillna("").astype(str)
    # Forma 1: AR_x: descripcion larga           (p.ej. "AR_1: Asistencial...")
    # Forma 2: AR_x nombre corto: descripcion larga (p.ej. "AR_1 Asistencial: ...")
    # Se acepta espacio(s) antes del ":" (p.ej. "AR_4 : descripcion")
    ar_pattern = re.compile(r"^AR_(\d+)(?:\s+[^:]+)?\s*:\s*.+", re.IGNORECASE)

    issues: list[str] = []
    prev_num: int | None = None
    found_any = False

    for c in disciplinar_cols:
        cell = ar_row.iloc[c].strip()
        if cell in _EMPTY or cell.upper() == "NAN":
            continue

        found_any = True
        col_letter = get_column_letter(c + 1)  # openpyxl es 1-based
        m = ar_pattern.match(cell)
        if not m:
            issues.append(
                f"Fila 2, columna {col_letter}: se esperaba el formato "
                f"'AR_x: descripcion' o 'AR_x nombre: descripcion', "
                f"pero se encontró '{cell}'."
            )
            continue

        num = int(m.group(1))
        if prev_num is None:
            if num != 1:
                issues.append(
                    f"Fila 2, columna {col_letter}: el primer Ámbito de Realización "
                    f"bajo 'ÁREA DE FORMACIÓN DISCIPLINAR' debe ser 'AR_1', "
                    f"pero se encontró 'AR_{num}'."
                )
        elif num < prev_num:
            issues.append(
                f"Fila 2, columna {col_letter}: el número de AR no es correlativo "
                f"— se encontró AR_{num} después de AR_{prev_num} "
                "(el número no puede decrecer)."
            )
        elif num > prev_num + 1:
            issues.append(
                f"Fila 2, columna {col_letter}: el número de AR no es correlativo "
                f"— se saltó de AR_{prev_num} a AR_{num} "
                f"(falta AR_{prev_num + 1})."
            )
        prev_num = num

    if not found_any:
        issues.append(
            "Fila 2: no se encontró ninguna celda con formato 'AR_x: descripcion' "
            "o 'AR_x nombre: descripcion' bajo 'ÁREA DE FORMACIÓN DISCIPLINAR'."
        )

    return issues


def _check_resultados_aprendizaje(df: pd.DataFrame) -> list[str]:
    """Check 5: fila 3 (row 2) — formato y correlatividad de los RA bajo DISCIPLINAR.

    Para las columnas cuyo encabezado en fila 1 (row 0) corresponde a
    ``'ÁREA DE FORMACIÓN DISCIPLINAR'`` (determinado por forward-fill), cada
    celda no vacía debe cumplir **todas** estas reglas:

    1. **Formato**: ``AR_x – RAy: <descripción larga>``  (guion medio, largo o corto).
    2. **Consistencia de AR**: el número ``x`` de ``AR_x`` debe coincidir con el
       número del ``AR_x`` de la celda directamente arriba en la fila 2 (fila de
       Ámbitos de Realización, después de forward-fill).
    3. **Correlatividad de RA**: los valores de ``y`` deben ser un entero
       correlativo global de izquierda a derecha bajo DISCIPLINAR:
       - El primer ``y`` encontrado debe ser ``1``.
       - Cada celda posterior puede repetir el número anterior (celda fusionada)
         o avanzar exactamente en 1 (``y == prev_y`` ó ``y == prev_y + 1``).

    Para las columnas fuera de DISCIPLINAR se conserva la comprobación mínima:
    al menos una celda debe contener ``"–"`` o ``":"``.
    """
    import re
    from openpyxl.utils import get_column_letter

    if df.shape[0] < 3 or df.shape[1] < 4:
        return []

    # Forward-fill fila 0 para identificar a qué área pertenece cada columna
    area_row_ffill = df.iloc[0].ffill().fillna("").astype(str)
    disciplinar_upper = "ÁREA DE FORMACIÓN DISCIPLINAR".upper()
    disciplinar_cols = [
        c for c in range(3, df.shape[1])
        if area_row_ffill.iloc[c].strip().upper() == disciplinar_upper
    ]

    # Fila 2 después de forward-fill para leer el AR de cada columna
    ar_row_ffill = df.iloc[1].ffill().fillna("").astype(str)
    # Fila 3 SIN forward-fill: cada RA es una celda independiente
    ra_row = df.iloc[2].fillna("").astype(str)

    # Patrón: AR_?x [–|—|-|<espacio>] R.?A.? y [: | espacio] descripción
    # Variantes aceptadas (todas observadas en matrices reales):
    #   AR_4 – RA11: desc          (estándar)
    #   AR_4 – RA 11: desc         (espacio entre RA y número)
    #   AR4 – RA11: desc           (sin guion bajo en AR)
    #   AR_4 - R.A.10: desc        (con puntos en RA, guion corto)
    #   AR_5 R.A. 15: desc         (sin separador, notación R.A. con espacio)
    #   AR_1 – RA2 Utiliza...      (sin ':' tras el número de RA)
    _RA_PATTERN = re.compile(
        r"^AR_?(\d+)\s*(?:[–—\-]\s*)?R\.?A\.?\s*(\d+)\s*[:\s].+",
        re.IGNORECASE | re.DOTALL,
    )
    # Extractor del número de AR desde la celda de fila 2 (acepta AR4 y AR_4)
    _AR_NUM_PATTERN = re.compile(r"^AR_?(\d+)", re.IGNORECASE)

    issues: list[str] = []

    # -----------------------------------------------------------------------
    # Validación estricta: columnas bajo DISCIPLINAR
    # -----------------------------------------------------------------------
    if disciplinar_cols:
        prev_ra_num: int | None = None
        found_any_disciplinar = False

        for c in disciplinar_cols:
            cell = ra_row.iloc[c].strip()
            if not cell or cell in _EMPTY or cell.upper() == "NAN":
                continue

            found_any_disciplinar = True
            col_letter = get_column_letter(c + 1)  # openpyxl es 1-based

            m = _RA_PATTERN.match(cell)
            if not m:
                issues.append(
                    f"Fila 3, columna {col_letter}: se esperaba el formato "
                    f"'AR_x – RAy: descripción' (también se aceptan variantes "
                    f"AR_?x, R.A.y, guion corto/largo u omitido, ':' opcional), "
                    f"pero se encontró '{cell}'."
                )
                continue

            ra_ar_num = int(m.group(1))   # x  (número de AR en la celda RA)
            ra_num    = int(m.group(2))   # y  (número correlativo de RA)

            # Regla 2: x debe coincidir con el AR de la fila superior
            ar_cell = ar_row_ffill.iloc[c].strip()
            ar_m = _AR_NUM_PATTERN.match(ar_cell)
            if ar_m:
                expected_ar = int(ar_m.group(1))
                if ra_ar_num != expected_ar:
                    issues.append(
                        f"Fila 3, columna {col_letter}: el número de AR en el RA "
                        f"(AR_{ra_ar_num}) no coincide con el AR de la celda "
                        f"superior en fila 2 (AR_{expected_ar}). "
                        f"Celda encontrada: '{cell}'."
                    )

            # Regla 3: y debe ser correlativo
            if prev_ra_num is None:
                if ra_num != 1:
                    issues.append(
                        f"Fila 3, columna {col_letter}: el primer Resultado de "
                        f"Aprendizaje bajo 'ÁREA DE FORMACIÓN DISCIPLINAR' debe "
                        f"ser RA1, pero se encontró RA{ra_num}."
                    )
            elif ra_num < prev_ra_num:
                issues.append(
                    f"Fila 3, columna {col_letter}: el número de RA no es "
                    f"correlativo — se encontró RA{ra_num} después de "
                    f"RA{prev_ra_num} (el número no puede decrecer)."
                )
            elif ra_num > prev_ra_num + 1:
                issues.append(
                    f"Fila 3, columna {col_letter}: el número de RA no es "
                    f"correlativo — se saltó de RA{prev_ra_num} a RA{ra_num} "
                    f"(falta RA{prev_ra_num + 1})."
                )
            prev_ra_num = ra_num

        if not found_any_disciplinar:
            issues.append(
                "Fila 3: no se encontró ninguna celda con formato "
                "'AR_x – RAy: descripción' bajo 'ÁREA DE FORMACIÓN DISCIPLINAR'."
            )

    # -----------------------------------------------------------------------
    # Verificación mínima global (detecta archivos completamente rotos)
    # -----------------------------------------------------------------------
    ra_any_ok = any(
        ("–" in cell or "—" in cell or ":" in cell)
        for cell in ra_row
        if cell.strip() not in _EMPTY
    )
    if not ra_any_ok:
        sample = [c for c in ra_row if c.strip() not in _EMPTY][:5]
        issues.append(
            "Fila 3 (Resultados de Aprendizaje): ninguna celda en toda la fila "
            "contiene un separador '–', '—' o ':'. "
            f"Valores encontrados: {sample}"
        )

    return issues


def _check_ra_basica(df: pd.DataFrame) -> list[str]:
    """Check 5b: fila 3 (row 2) — formato y correlatividad de los RA bajo BÁSICA.

    Para las columnas cuyo encabezado en fila 1 (row 0) corresponde a
    ``'ÁREA DE FORMACIÓN BÁSICA'`` (determinado por forward-fill), cada
    celda no vacía debe cumplir todas estas reglas:

    1. **Formato**: ``FB_RAx: descripción`` o
       ``FB_RAx: descripción corta: descripción larga``.
    2. **Correlatividad**: los valores de ``x`` deben ser un entero correlativo
       de izquierda a derecha: el primero debe ser ``1``, y cada celda posterior
       puede repetir el número anterior (celda fusionada) o avanzar en 1.
    """
    import re
    from openpyxl.utils import get_column_letter

    if df.shape[0] < 3 or df.shape[1] < 4:
        return []

    area_row_ffill = df.iloc[0].ffill().fillna("").astype(str)
    basica_upper = "ÁREA DE FORMACIÓN BÁSICA".upper()
    basica_cols = [
        c for c in range(3, df.shape[1])
        if area_row_ffill.iloc[c].strip().upper() == basica_upper
    ]

    if not basica_cols:
        return []  # sección opcional — su ausencia ya la valida _check_area_title_order

    ra_row = df.iloc[2].fillna("").astype(str)
    _FB_RA_PATTERN = re.compile(r"^FB_RA(\d+)\s*:\s*.+", re.IGNORECASE | re.DOTALL)

    issues: list[str] = []
    prev_num: int | None = None
    found_any = False

    for c in basica_cols:
        cell = ra_row.iloc[c].strip()
        if not cell or cell in _EMPTY or cell.upper() == "NAN":
            continue

        found_any = True
        col_letter = get_column_letter(c + 1)
        m = _FB_RA_PATTERN.match(cell)
        if not m:
            issues.append(
                f"Fila 3, columna {col_letter}: bajo 'ÁREA DE FORMACIÓN BÁSICA' "
                f"se esperaba el formato 'FB_RAx: descripción' "
                f"(opcionalmente 'FB_RAx: desc_corta: desc_larga'), "
                f"pero se encontró '{cell}'."
            )
            continue

        num = int(m.group(1))
        if prev_num is None:
            if num != 1:
                issues.append(
                    f"Fila 3, columna {col_letter}: el primer RA bajo "
                    f"'ÁREA DE FORMACIÓN BÁSICA' debe ser 'FB_RA1', "
                    f"pero se encontró 'FB_RA{num}'."
                )
        elif num < prev_num:
            issues.append(
                f"Fila 3, columna {col_letter}: el número de FB_RA no es correlativo "
                f"— se encontró FB_RA{num} después de FB_RA{prev_num} "
                "(el número no puede decrecer)."
            )
        elif num > prev_num + 1:
            issues.append(
                f"Fila 3, columna {col_letter}: el número de FB_RA no es correlativo "
                f"— se saltó de FB_RA{prev_num} a FB_RA{num} "
                f"(falta FB_RA{prev_num + 1})."
            )
        prev_num = num

    if not found_any:
        issues.append(
            "Fila 3: no se encontró ninguna celda con formato 'FB_RAx: descripción' "
            "bajo 'ÁREA DE FORMACIÓN BÁSICA'."
        )

    return issues


def _check_ra_general(df: pd.DataFrame) -> list[str]:
    """Check 5c: fila 3 (row 2) — formato y correlatividad de los RA bajo GENERAL / EJE.

    Para las columnas cuyo encabezado en fila 1 (row 0) corresponde a
    ``'ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN INTEGRAL'`` o a su alias
    ``'HABILIDADES TRANSVERSALES'`` (determinado por forward-fill), cada
    celda no vacía debe cumplir todas estas reglas, **excepto** las celdas
    cuyo valor sea literalmente ``"HABILIDADES TRANSVERSALES"`` (que actúan
    como sub-encabezado y se ignoran):

    1. **Formato**: ``FG_RAx: descripción`` o
       ``FG_RAx: descripción corta: descripción larga``.
    2. **Correlatividad**: los valores de ``x`` deben ser un entero correlativo
       de izquierda a derecha: el primero debe ser ``1``, y cada celda posterior
       puede repetir el número anterior (celda fusionada) o avanzar en 1.
    """
    import re
    from openpyxl.utils import get_column_letter

    if df.shape[0] < 3 or df.shape[1] < 4:
        return []

    area_row_ffill = df.iloc[0].ffill().fillna("").astype(str)
    general_upper = "ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN INTEGRAL".upper()
    habilidades_upper = "HABILIDADES TRANSVERSALES".upper()
    general_cols = [
        c for c in range(3, df.shape[1])
        if area_row_ffill.iloc[c].strip().upper() in (general_upper, habilidades_upper)
    ]

    if not general_cols:
        return []  # sección requerida — su ausencia ya la valida _check_area_title_order

    ra_row = df.iloc[2].fillna("").astype(str)
    _FG_RA_PATTERN = re.compile(r"^FG_RA(\d+)\s*:\s*.+", re.IGNORECASE | re.DOTALL)

    issues: list[str] = []
    prev_num: int | None = None
    found_any = False

    for c in general_cols:
        cell = ra_row.iloc[c].strip()
        if not cell or cell in _EMPTY or cell.upper() == "NAN":
            continue
        # sub-encabezado permitido — se omite de la validación de formato
        if cell.strip().upper() == habilidades_upper:
            continue

        found_any = True
        col_letter = get_column_letter(c + 1)
        m = _FG_RA_PATTERN.match(cell)
        if not m:
            issues.append(
                f"Fila 3, columna {col_letter}: bajo 'ÁREA DE FORMACIÓN GENERAL / "
                f"EJE DE FORMACIÓN INTEGRAL' se esperaba el formato "
                f"'FG_RAx: descripción' "
                f"(opcionalmente 'FG_RAx: desc_corta: desc_larga'), "
                f"pero se encontró '{cell}'."
            )
            continue

        num = int(m.group(1))
        if prev_num is None:
            if num != 1:
                issues.append(
                    f"Fila 3, columna {col_letter}: el primer RA bajo "
                    f"'ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN INTEGRAL' "
                    f"debe ser 'FG_RA1', pero se encontró 'FG_RA{num}'."
                )
        elif num < prev_num:
            issues.append(
                f"Fila 3, columna {col_letter}: el número de FG_RA no es correlativo "
                f"— se encontró FG_RA{num} después de FG_RA{prev_num} "
                "(el número no puede decrecer)."
            )
        elif num > prev_num + 1:
            issues.append(
                f"Fila 3, columna {col_letter}: el número de FG_RA no es correlativo "
                f"— se saltó de FG_RA{prev_num} a FG_RA{num} "
                f"(falta FG_RA{prev_num + 1})."
            )
        prev_num = num

    if not found_any:
        issues.append(
            "Fila 3: no se encontró ninguna celda con formato 'FG_RAx: descripción' "
            "bajo 'ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN INTEGRAL' "
            "(excluyendo celdas con valor 'HABILIDADES TRANSVERSALES')."
        )

    return issues


def _check_niveles_logro(df: pd.DataFrame) -> list[str]:
    """Check 6: fila 4 (row 3) — formato y ciclo correlativo de niveles de logro.

    Para las columnas bajo ``'ÁREA DE FORMACIÓN DISCIPLINAR'``,
    ``'ÁREA DE FORMACIÓN BÁSICA'`` y
    ``'ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN INTEGRAL'``
    (incluido su alias ``'HABILIDADES TRANSVERSALES'``), **excepto** las
    columnas donde la fila 3 (row 2) contiene el sub-encabezado
    ``'HABILIDADES TRANSVERSALES'``, cada celda no vacía debe:

    1. Coincidir con el patrón
       ``N. (INICIAL|INTERMEDIO|AVANZADO|TITULACION): descripción``.
    2. Seguir el ciclo correlativo INICIAL → INTERMEDIO → AVANZADO/TITULACION
       → INICIAL, donde cada celda puede repetir el nivel anterior (celda
       fusionada) o avanzar al siguiente en el ciclo.
    """
    import re
    from openpyxl.utils import get_column_letter

    if df.shape[0] < 4 or df.shape[1] < 4:
        return []

    area_row_ffill = df.iloc[0].ffill().fillna("").astype(str)
    ra_row_ffill = df.iloc[2].ffill().fillna("").astype(str)  # fila 3 (index 2): forward-filled para detectar celdas fusionadas

    disciplinar_upper = "ÁREA DE FORMACIÓN DISCIPLINAR".upper()
    basica_upper = "ÁREA DE FORMACIÓN BÁSICA".upper()
    general_upper = "ÁREA DE FORMACIÓN GENERAL / EJE DE FORMACIÓN INTEGRAL".upper()
    habilidades_upper = "HABILIDADES TRANSVERSALES".upper()

    eligible_areas = {disciplinar_upper, basica_upper, general_upper, habilidades_upper}

    eligible_cols = [
        c for c in range(3, df.shape[1])
        if area_row_ffill.iloc[c].strip().upper() in eligible_areas
        and ra_row_ffill.iloc[c].strip().upper() != habilidades_upper
    ]

    if not eligible_cols:
        return []

    lvl_row = df.iloc[3].fillna("").astype(str)

    _NIVEL_PATTERN = re.compile(
        r"^N\.\s*(INICIAL|INTERMEDIO|AVANZADO|TITULACI[OÓ]N)\s*:",
        re.IGNORECASE,
    )
    # Posición en el ciclo: 0=INICIAL, 1=INTERMEDIO, 2=AVANZADO/TITULACION
    _NIVEL_POS: dict[str, int] = {
        "INICIAL": 0,
        "INTERMEDIO": 1,
        "AVANZADO": 2,
        "TITULACION": 2,
        "TITULACIÓN": 2,
    }
    _NIVEL_LABEL = ["INICIAL", "INTERMEDIO", "AVANZADO"]

    issues: list[str] = []
    prev_pos: int | None = None
    found_any = False

    for c in eligible_cols:
        cell = lvl_row.iloc[c].strip()
        if not cell or cell in _EMPTY or cell.upper() == "NAN":
            continue

        found_any = True
        col_letter = get_column_letter(c + 1)
        m = _NIVEL_PATTERN.match(cell)
        if not m:
            issues.append(
                f"Fila 4, columna {col_letter}: se esperaba "
                f"'N. INICIAL: descripción', 'N. INTERMEDIO: descripción' o "
                f"'N. AVANZADO: descripción' (también 'N. TITULACION: descripción'), "
                f"pero se encontró '{cell}'."
            )
            continue

        level_raw = m.group(1).upper()
        # normalizar tilde para la búsqueda en el dict
        level_key = level_raw.replace("Ó", "O").replace("Á", "A")
        pos = _NIVEL_POS.get(level_key, _NIVEL_POS.get(level_raw, -1))
        if pos == -1:
            continue  # no debería ocurrir dado el patrón del regex

        if prev_pos is None:
            if pos != 0:
                issues.append(
                    f"Fila 4, columna {col_letter}: el primer nivel de logro debe ser "
                    f"'N. INICIAL', pero se encontró '{cell}'."
                )
        else:
            expected_next = (prev_pos + 1) % 3
            if pos != prev_pos and pos != expected_next:
                prev_label = _NIVEL_LABEL[prev_pos]
                exp_label = _NIVEL_LABEL[expected_next]
                issues.append(
                    f"Fila 4, columna {col_letter}: nivel de logro fuera del ciclo "
                    f"INICIAL→INTERMEDIO→AVANZADO — después de N. {prev_label} se "
                    f"esperaba N. {prev_label} (repetición) o N. {exp_label} "
                    f"(siguiente), pero se encontró '{cell}'."
                )
        prev_pos = pos

    if not found_any:
        sample_cols = [get_column_letter(c + 1) for c in eligible_cols[:5]]
        suffix = " ..." if len(eligible_cols) > 5 else ""
        issues.append(
            "Fila 4: no se encontró ninguna celda con formato "
            "'N. INICIAL/INTERMEDIO/AVANZADO: ...' en las columnas elegibles "
            f"({', '.join(sample_cols)}{suffix})."
        )

    return issues


def _check_nombres_asignatura(df: pd.DataFrame) -> list[str]:
    """Check 7: la columna C (índice 2), desde la fila 6 (row 5), tiene nombres de asignatura."""
    col_c = df.iloc[5:, 2].fillna("").astype(str)
    if not any(v.strip() not in _EMPTY for v in col_c):
        return [
            "Columna C (índice 2) desde la fila 6 en adelante: no se encontró "
            "ningún nombre de asignatura. El parser espera los nombres de asignatura "
            "exactamente en esa columna a partir de esa fila."
        ]
    return []


def _check_area_formacion(df: pd.DataFrame) -> list[str]:
    """Check 8: la columna A (índice 0), desde la fila 6 (row 5), tiene valores de área de formación."""
    col_a = df.iloc[5:, 0].fillna("").astype(str)
    if not any(v.strip() not in _EMPTY for v in col_a):
        return [
            "Columna A (índice 0) desde la fila 6: no se encontró ningún valor de "
            "'Área de formación'. El parser lee el área de formación de la columna A."
        ]
    return []


def _check_semestre_romano(df: pd.DataFrame) -> list[str]:
    """Check 9: la columna B (índice 1), desde la fila 6 (row 5), tiene numerales romanos de semestre."""
    col_b = df.iloc[5:, 1].fillna("").astype(str)
    romanos = [v.strip().upper() for v in col_b if v.strip().upper() in _ROMAN_NUMERALS]
    if not romanos:
        return [
            "Columna B (índice 1) desde la fila 6: no se encontró ningún numeral "
            f"romano de semestre. Se esperan valores como {sorted(_ROMAN_NUMERALS)}. "
            "El parser lee el semestre de la columna B y lo convierte con roman_to_int()."
        ]
    return []


def _check_tributacion_values(df: pd.DataFrame) -> list[str]:
    """Check 10: en las filas de asignatura (≥ fila 6), dentro de columnas de área válida, hay al menos un '1'."""
    area_row = df.iloc[0].ffill().fillna("").astype(str)
    area_titles_upper = [a.upper() for a in AREA_TITLES]
    valid_cols = [
        i for i in range(df.shape[1])
        if area_row.iloc[i].strip().upper() in area_titles_upper
    ]
    if not valid_cols:
        return []  # check 3 ya reportó la ausencia de áreas
    tributacion_body = df.iloc[5:, valid_cols].astype(str)
    if not (tributacion_body == TRIBUTACION_VALUE).any().any():
        return [
            f"No se encontró ninguna celda con el valor '{TRIBUTACION_VALUE}' en las "
            "columnas de área válida (filas ≥6). O las asignaturas no tienen tributación "
            "marcada, o el valor de tributación usa otro carácter (¿'x', '✓', '1.0'?)."
        ]
    return []


# ---------------------------------------------------------------------------
# Función pública
# ---------------------------------------------------------------------------

def validate_matrix_structure(
    xlsx_path: Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> list[str]:
    """Valida que el Excel tenga la estructura esperada por ``parse_matrix``.

    Ejecuta en orden los diez checks del contrato estructural. Cada check
    es una función privada independiente que devuelve una lista de mensajes
    de error (vacía si esa condición se cumple).

    Args:
        xlsx_path:  Ruta al archivo Excel de la Matriz de Tributación.
        sheet_name: Nombre de la hoja esperada (por defecto ``"Asignaturas - RA"``).

    Returns:
        Lista de mensajes de error (vacía si el archivo es válido).
        Cada mensaje describe exactamente qué está mal y en qué posición.

    Examples:
        >>> issues = validate_matrix_structure(Path("Matriz.xlsx"))
        >>> assert issues == [], "\\n".join(issues)
    """
    xlsx_path = _resolve_path(Path(xlsx_path))

    # Checks previos a la carga del DataFrame — fallan rápido si no hay archivo u hoja
    issues = _check_file_exists(xlsx_path)
    if issues:
        return issues

    issues = _check_sheet_exists(xlsx_path, sheet_name)
    if issues:
        return issues

    df = pd.read_excel(str(xlsx_path), sheet_name=sheet_name, header=None)

    # Check de dimensiones — si falla, los demás checks no tienen sentido
    issues = _check_dimensions(df)
    if issues:
        return issues

    # Check de cabeceras fijas en columna C — falla rápido si la estructura no es la esperada
    issues = _check_col_c_headers(df)
    if issues:
        return issues

    # Checks independientes sobre el contenido de la hoja
    issues = []
    issues += _check_col_a_header(df)
    issues += _check_col_b_header(df)
    issues += _check_area_title_order(df)
    issues += _check_area_row_gaps(xlsx_path, sheet_name)
    issues += _check_ambitos_realizacion(df)
    issues += _check_resultados_aprendizaje(df)
    issues += _check_ra_basica(df)
    issues += _check_ra_general(df)
    issues += _check_niveles_logro(df)
    issues += _check_nombres_asignatura(df)
    issues += _check_area_formacion(df)
    issues += _check_semestre_romano(df)
    issues += _check_tributacion_values(df)

    return issues
